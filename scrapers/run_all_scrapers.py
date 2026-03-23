"""
run_all_scrapers.py
───────────────────
Master runner — scrapes BOTH sites and saves:
  • Individual files per site  (CSV + JSON + XLSX)
  • A combined merged file     (CSV + JSON + XLSX)

Usage:
    python run_all_scrapers.py

Requirements:
    pip install requests openpyxl
"""

import subprocess
import sys
import json
import csv
import os
import time
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

OUTPUT_DIR = "output"
TIMESTAMP  = datetime.now().strftime("%Y%m%d_%H%M%S")


# ── Inline scraper (same logic as individual files, combined here) ─────────────

import requests
import re


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


def clean_html(html: str) -> str:
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"&nbsp;|&amp;|&lt;|&gt;", lambda m: {
        "&nbsp;": " ", "&amp;": "&", "&lt;": "<", "&gt;": ">"
    }.get(m.group(), m.group()), text)
    return re.sub(r"\s+", " ", text).strip()


def fetch_shopify(base_url: str, site_label: str, extra_fields_fn=None) -> list[dict]:
    api_url  = f"{base_url}/products.json"
    all_rows = []
    page_info = None
    page_num  = 0

    print(f"\n{'─'*60}")
    print(f"  Fetching: {site_label}")
    print(f"  URL     : {api_url}")
    print(f"{'─'*60}")

    while True:
        page_num += 1
        params = {"limit": 250}
        if page_info:
            params["page_info"] = page_info

        try:
            resp = requests.get(api_url, headers=HEADERS, params=params, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"  [ERROR] Page {page_num}: {e}")
            break

        products = resp.json().get("products", [])
        if not products:
            print(f"  No more products at page {page_num}.")
            break

        # Cursor for next page
        link = resp.headers.get("Link", "")
        next_cursor = None
        if 'rel="next"' in link:
            m = re.search(r'page_info=([^&>]+)[^>]*>;\s*rel="next"', link)
            if m:
                next_cursor = m.group(1)

        for product in products:
            rows = parse_shopify_product(product, base_url, site_label, extra_fields_fn)
            all_rows.extend(rows)

        print(f"  Page {page_num:>3}: {len(products):>4} products  |  Total rows: {len(all_rows)}")

        if not next_cursor:
            break
        page_info = next_cursor
        time.sleep(0.5)

    print(f"  ✓ Finished. {len(all_rows)} total rows collected.")
    return all_rows


def parse_shopify_product(product: dict, base_url: str, site_label: str,
                           extra_fields_fn=None) -> list[dict]:
    pid          = product.get("id", "")
    title        = product.get("title", "")
    description  = clean_html(product.get("body_html", ""))
    vendor       = product.get("vendor", "")
    product_type = product.get("product_type", "")
    tags_list    = product.get("tags", [])
    tags         = ", ".join(tags_list)
    handle       = product.get("handle", "")
    product_url  = f"{base_url}/products/{handle}"
    published_at = product.get("published_at", "")

    image_urls = [i["src"] for i in product.get("images", []) if i.get("src")]
    main_image = image_urls[0] if image_urls else ""
    all_images = " | ".join(image_urls)

    variants = product.get("variants", [{}])
    rows = []

    for variant in variants:
        sku           = variant.get("sku", "")
        price         = variant.get("price", "")
        compare_price = variant.get("compare_at_price", "") or ""
        vtitle        = variant.get("title", "")
        inv_qty       = variant.get("inventory_quantity", None)
        inv_policy    = variant.get("inventory_policy", "")

        if inv_qty is None:
            stock = "Unknown"
        elif inv_qty > 0:
            stock = "In Stock"
        elif inv_policy == "continue":
            stock = "Available (Backorder)"
        else:
            stock = "Out of Stock"

        row = {
            "Site"              : site_label,
            "Source URL"        : base_url,
            "Product ID"        : pid,
            "Product URL"       : product_url,
            "Name"              : title,
            "Variant"           : vtitle if vtitle != "Default Title" else "",
            "SKU"               : sku,
            "Category"          : product_type,
            "Tags"              : tags,
            "Vendor / Brand"    : vendor,
            "Price (GBP)"       : price,
            "Compare At Price"  : compare_price,
            "Stock Status"      : stock,
            "Inventory Qty"     : inv_qty if inv_qty is not None else "",
            "Description"       : description,
            "Main Image URL"    : main_image,
            "All Image URLs"    : all_images,
            "Published At"      : published_at,
        }

        # Site-specific extras
        if extra_fields_fn:
            row.update(extra_fields_fn(product, variant, tags_list))

        rows.append(row)

    return rows


def ukfd_extras(product: dict, variant: dict, tags_list: list) -> dict:
    """Extra fields specific to UK Flooring Direct."""
    m2_tag = next((t for t in tags_list
                   if "per-m2" in t.lower() or "per m2" in t.lower()), "")
    return {"Price Per m2 Tag": m2_tag}


# ── Output writers ─────────────────────────────────────────────────────────────

def save_csv(rows: list[dict], filepath: str, label: str):
    if not rows:
        return
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"  [{label}] CSV  → {filepath}")


def save_json(rows: list[dict], filepath: str, label: str):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"  [{label}] JSON → {filepath}")


def save_xlsx(rows: list[dict], filepath: str, label: str,
              header_color: str = "1F3864"):
    if not XLSX_AVAILABLE or not rows:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fieldnames = list(rows[0].keys())
    for ci, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=ci, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border

    for ri, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="F2F2F2") if ri % 2 == 0 else None
        for ci, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(field, ""))
            cell.border = border
            if fill:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"  [{label}] XLSX → {filepath}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    start = time.time()

    print("\n" + "═"*60)
    print("  PRODUCT SCRAPER — Two Shopify Stores")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("═"*60)

    # ── Site 1: HDEW Cameras ───────────────────────────────────────
    hdew_rows = fetch_shopify(
        base_url   = "https://hdewcameras.co.uk",
        site_label = "HDEW Cameras",
    )
    base1 = f"{OUTPUT_DIR}/hdew_cameras_{TIMESTAMP}"
    save_csv(hdew_rows,  f"{base1}.csv",  "HDEW")
    save_json(hdew_rows, f"{base1}.json", "HDEW")
    save_xlsx(hdew_rows, f"{base1}.xlsx", "HDEW", header_color="1F3864")

    # ── Site 2: UK Flooring Direct ─────────────────────────────────
    ukfd_rows = fetch_shopify(
        base_url       = "https://www.ukflooringdirect.co.uk",
        site_label     = "UK Flooring Direct",
        extra_fields_fn= ukfd_extras,
    )
    base2 = f"{OUTPUT_DIR}/ukflooring_direct_{TIMESTAMP}"
    save_csv(ukfd_rows,  f"{base2}.csv",  "UKFD")
    save_json(ukfd_rows, f"{base2}.json", "UKFD")
    save_xlsx(ukfd_rows, f"{base2}.xlsx", "UKFD", header_color="1E5631")

    # ── Combined ───────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print("  Merging both sites into combined output…")
    combined = hdew_rows + ukfd_rows
    base3 = f"{OUTPUT_DIR}/combined_all_products_{TIMESTAMP}"
    save_csv(combined,  f"{base3}.csv",  "COMBINED")
    save_json(combined, f"{base3}.json", "COMBINED")
    save_xlsx(combined, f"{base3}.xlsx", "COMBINED", header_color="4A235A")

    # ── Summary ────────────────────────────────────────────────────
    elapsed = time.time() - start
    print(f"\n{'═'*60}")
    print("  SUMMARY")
    print(f"{'─'*60}")
    print(f"  HDEW Cameras         : {len(hdew_rows):>6} rows")
    print(f"  UK Flooring Direct   : {len(ukfd_rows):>6} rows")
    print(f"  Combined             : {len(combined):>6} rows")
    print(f"  Time elapsed         : {elapsed:.1f}s")
    print(f"  Output folder        : {os.path.abspath(OUTPUT_DIR)}/")
    print(f"{'═'*60}")
    print("  ✅ All done!\n")


if __name__ == "__main__":
    main()
