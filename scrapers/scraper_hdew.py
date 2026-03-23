"""
HDEW Cameras - Shopify Product Scraper
Site: https://hdewcameras.co.uk
Method: Shopify /products.json API (no auth required)

Fields extracted:
  - Name, Description, Price, Images, SKU, Category, Stock Status

Output: products_hdew.csv / .json / .xlsx
"""

import requests
import json
import csv
import time
import re
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("[WARNING] openpyxl not installed. Excel output will be skipped.")
    print("          Run: pip install openpyxl")

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL    = "https://hdewcameras.co.uk"
API_URL     = f"{BASE_URL}/products.json"
LIMIT       = 250          # Max per page Shopify allows
DELAY       = 0.5          # Seconds between requests (be polite)
OUTPUT_DIR  = "output"
SITE_NAME   = "hdew_cameras"
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


def clean_html(html: str) -> str:
    """Strip HTML tags and clean whitespace from description."""
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"&amp;", "&", text)
    text = re.sub(r"&lt;", "<", text)
    text = re.sub(r"&gt;", ">", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def fetch_all_products() -> list[dict]:
    """
    Fetch every product using Shopify cursor-based pagination.
    Returns a flat list of normalised product dicts.
    """
    products = []
    page_info = None
    page_num  = 0

    print(f"[HDEW] Starting fetch from {API_URL}")

    while True:
        page_num += 1
        params = {"limit": LIMIT}
        if page_info:
            params["page_info"] = page_info

        try:
            resp = requests.get(API_URL, headers=HEADERS, params=params, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"[HDEW] Request error on page {page_num}: {e}")
            break

        data = resp.json().get("products", [])
        if not data:
            print(f"[HDEW] No more products at page {page_num}. Done.")
            break

        # Parse Link header for next page cursor
        link_header = resp.headers.get("Link", "")
        next_page_info = None
        if 'rel="next"' in link_header:
            match = re.search(r'page_info=([^&>]+)[^>]*>;\s*rel="next"', link_header)
            if match:
                next_page_info = match.group(1)

        for product in data:
            rows = parse_product(product)
            products.extend(rows)

        print(f"[HDEW] Page {page_num}: fetched {len(data)} products "
              f"(total rows so far: {len(products)})")

        if not next_page_info:
            print("[HDEW] No next page link. Finished pagination.")
            break

        page_info = next_page_info
        time.sleep(DELAY)

    return products


def parse_product(product: dict) -> list[dict]:
    """
    Each Shopify product can have multiple variants.
    We create one row per variant so every SKU is captured.
    """
    rows = []

    product_id   = product.get("id", "")
    title        = product.get("title", "")
    description  = clean_html(product.get("body_html", ""))
    vendor       = product.get("vendor", "")
    product_type = product.get("product_type", "")
    tags         = ", ".join(product.get("tags", []))
    handle       = product.get("handle", "")
    product_url  = f"{BASE_URL}/products/{handle}"
    published_at = product.get("published_at", "")

    # All image URLs for this product
    image_urls = [img.get("src", "") for img in product.get("images", []) if img.get("src")]
    images_str = " | ".join(image_urls)
    main_image = image_urls[0] if image_urls else ""

    variants = product.get("variants", [])
    if not variants:
        variants = [{}]  # Ensure at least one row even with no variants

    for variant in variants:
        sku             = variant.get("sku", "")
        price           = variant.get("price", "")
        compare_price   = variant.get("compare_at_price", "")
        variant_title   = variant.get("title", "")
        inventory_qty   = variant.get("inventory_quantity", None)
        inventory_policy= variant.get("inventory_management", "")

        # Stock status logic
        if inventory_qty is None:
            stock_status = "Unknown"
        elif inventory_qty > 0:
            stock_status = "In Stock"
        elif variant.get("inventory_policy") == "continue":
            stock_status = "Available (Backorder)"
        else:
            stock_status = "Out of Stock"

        rows.append({
            "Source Site"       : BASE_URL,
            "Product ID"        : product_id,
            "Product URL"       : product_url,
            "Name"              : title,
            "Variant"           : variant_title if variant_title != "Default Title" else "",
            "SKU"               : sku,
            "Category"          : product_type,
            "Tags"              : tags,
            "Vendor / Brand"    : vendor,
            "Price (GBP)"       : price,
            "Compare At Price"  : compare_price if compare_price else "",
            "Stock Status"      : stock_status,
            "Inventory Qty"     : inventory_qty if inventory_qty is not None else "",
            "Description"       : description,
            "Main Image URL"    : main_image,
            "All Image URLs"    : images_str,
            "Published At"      : published_at,
        })

    return rows


def save_csv(products: list[dict], filepath: str):
    if not products:
        print("[HDEW] No data to save.")
        return
    fieldnames = list(products[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(products)
    print(f"[HDEW] CSV saved  → {filepath}  ({len(products)} rows)")


def save_json(products: list[dict], filepath: str):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"[HDEW] JSON saved → {filepath}  ({len(products)} records)")


def save_xlsx(products: list[dict], filepath: str):
    if not XLSX_AVAILABLE:
        return
    if not products:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header style
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    fieldnames = list(products[0].keys())
    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, product in enumerate(products, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            value = product.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto column width (capped at 60)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"[HDEW] XLSX saved → {filepath}  ({len(products)} rows)")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{OUTPUT_DIR}/{SITE_NAME}_{timestamp}"

    products = fetch_all_products()

    if not products:
        print("[HDEW] No products found. Check the site or your network.")
        return

    print(f"\n[HDEW] Total rows collected: {len(products)}")

    save_csv(products,  f"{base_name}.csv")
    save_json(products, f"{base_name}.json")
    save_xlsx(products, f"{base_name}.xlsx")

    print(f"\n[HDEW] ✅ Done! Files saved in '{OUTPUT_DIR}/' folder.")


if __name__ == "__main__":
    main()
